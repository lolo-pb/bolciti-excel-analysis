from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def style_financial_sheet(
    filepath: str,
    sheet_name: str = "Resumen",
    label_col: int = 1,
    header_row: int = 1,
    money_format: str = '"$"#,##0.00',
    freeze_panes_cell: str = "B2",
    special_row_fills: dict[str, str] | None = None,
) -> None:
    """
    Apply Excel styling to a report-like sheet:
    - header style
    - borders
    - number formatting for numeric cells
    - highlights special rows by label (first column)
    - auto column widths
    - freeze panes
    """

    if special_row_fills is None:
        # default highlights (you can override from main)
        special_row_fills = {
            "TOTAL_SPENDINGS": "FFF2CC",  # light yellow
            "FACTURACION": "BDD7EE",      # light blue
            "RESULTADO": "E2EFDA",        # light green
        }

    wb = load_workbook(filepath)
    ws = wb[sheet_name]

    # --- styles
    zebra_fill = PatternFill("solid", fgColor="02D4C6")
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    label_gray = PatternFill("solid", fgColor="E6E6E6")        
    label_gray_dark = PatternFill("solid", fgColor="CFCFCF")  
    darker_labels = {"confeccion","impresion","extrusion","echado","oficina","gral",}
    bold = Font(bold=True)

    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # freeze panes (nice UX)
    if freeze_panes_cell:
        ws.freeze_panes = freeze_panes_cell

    # Determine used range
    max_row = ws.max_row
    max_col = ws.max_column

    # --- header row formatting
    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = bold
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    # --- cell borders + alignment + number formats
    for r in range(header_row + 1, max_row + 1):
        # label cell
        label_cell = ws.cell(row=r, column=label_col)
        label_cell.border = thin_border
        label_cell.alignment = left

        # numeric cells: everything except label column
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border

            if c == label_col:
                continue

            # Only apply money format to numeric cells
            # (if it’s text, openpyxl will ignore number_format visually)
            cell.number_format = money_format
            cell.alignment = right

    # --- zebra striping (even rows only, excluding header)
    for r in range(header_row + 1, max_row + 1):
        # Excel rows: apply to even-numbered data rows
        if (r - header_row) % 2 == 0:
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = zebra_fill

    # --- gray background for label column (index / seccion)
    for r in range(header_row + 1, max_row + 1):
        cell = ws.cell(row=r, column=label_col)
        label = cell.value
    
        if not isinstance(label, str):
            continue
        
        label_key = label.strip().lower()
    
        # darker gray for selected sections
        if label_key in darker_labels:
            cell.fill = label_gray_dark
            cell.font = Font(bold=True)
        else:
            cell.fill = label_gray
    
    # --- highlight special rows by label text (first column)
    for r in range(header_row + 1, max_row + 1):
        label = ws.cell(row=r, column=label_col).value
        if not isinstance(label, str):
            continue

        label_key = label.strip()
        if label_key in special_row_fills:
            fill = PatternFill("solid", fgColor=special_row_fills[label_key])
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill
                cell.font = bold

    # --- optional: color negative numbers red (common finance convention)
    # keep it simple: just set font color when cell value is numeric and < 0
    for r in range(header_row + 1, max_row + 1):
        for c in range(label_col + 1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if isinstance(v, (int, float)) and v < 0:
                cell.font = Font(color="9C0006", bold=cell.font.bold if cell.font else False)

    # --- auto column widths
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    wb.save(filepath)
